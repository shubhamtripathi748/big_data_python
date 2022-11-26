# print("Hello World Shubham")

patient_name = 'John Smith'
patient_age = 20
patient_is_new = True

# person_name = input('please enter your name ? ')
# person_like_color = input('which colour you like ? ')
# print(person_name+" likes "+person_like_color)

# person_weight = input("enter your weight(in pounds ) ? ")

# pounds_to_kg = float(person_weight)*0.45359237
# print(pounds_to_kg)

string_index = 'Hello Shubham for Python'

# print(string_index[-1])
# print(string_index[0])

# print(string_index[0:5])
# print(string_index[1:-1])
# copy_str = string_index[:]
# print(copy_str)


first = 'Shubham'
last = 'Tripathi'
formattedstr = f'{first} {last} is a coder'
# print(formattedstr)
title = 'Python for Beginner'

# print(title.upper())

# print(title.lower())
# print(title.find('B'))
# print('for' in title)

# is_hot = True
#
# if is_hot:
#     print(" it's hot day!!")
# else:
#     print(" it's not hot!!")
# print('Enjoy your day ')

# house_price = 1000000
# buyer_credit = True
# down_payment=0
# if buyer_credit :
#     down_payment = house_price*10/100
# else:
#     down_payment = house_price * 20 / 100
# print(f"down payment ${down_payment}")

# logical operator
# and or not
#
# applicant_income=True
# applicant_credit=True
#
# if applicant_income and applicant_credit :
#     print("eligible for loan")
#
# name='SHUBHAM'
# name_len = len(name)
# if  name_len< 3 :
#     print('name  must be more then 3 char')
# elif name_len> 50 :
#     print('name  must be less then 50 char')
# else :
#     print("name is good ",name)

#
# weight=int(input('Enter your weight '))
# unit=input('K or L ')
#
# if unit.upper() == 'L':
#     converted_w = weight*.45
#     print(f"your weight {converted_w} kg")
# elif unit.upper() == 'K':
#     converted_w = weight / .45
#     print(f"your weight {converted_w} L ")

# while loop example
#
# x=1
#
# y=10
# while x<=3:
#     guess_number=int(input("Guess the Number "))
#     x += 1
#     if guess_number == y:
#         print(f"number you have enter is matched")
#         break
# else :
#    print(f"number you have enter is not matched pls try again")


# CAR EXAMPLE
#
#
# is_continue = True
# is_start=False
# is_stop=False
# while is_continue:
#
#     car_command = input('add command ').lower()
#     if car_command == 'start' :
#         if is_start :
#             print('car already started')
#         else :
#           print('car started ready to go!!')
#           is_start = True
#     elif car_command == 'stop':
#         if not is_start :
#             print('car not started')
#         if is_stop :
#             print('car already stopped')
#         else :
#            print('car stopped')
#            is_stop=True
#     elif car_command == 'quit':
#         break
#     else:
#         print('i dont understand this command',car_command)


# for loop example
#
# for item in range(1,10,2):
#     print(item)

# total price in cart
# total_value=0
# for item in range(1,10):
#     total_value+=item
# print("total_value ",total_value)


# nested loop
#
# for x in range(5):
#     for y in range(5):
#         print(f'({x}, {y})')
#
# number=[5,2,5,2,2]
# for x in number:
#         str=''
#         for y in range(x):
#             str+='x';
#         print(str)


# find max number in the list
#
# number=[10,23,13,99,5]
#
# max = number[0]
# for item in number:
#     if max<item:
#         max=item
#
# print(max);

# list
#
# number=[10,23,5,13,99,5,10]
#
# not_duplicate=[]
#
# for item in number:
#     if number.count(item) > 1 :
#         number.remove(item)
# print(number)

# tuples is immutable and we can not change it

# tuples=(10,2,3,35)
# print(tuples[1])
#
# x,y,z=tuples
# print(x,y,z)


# dict
#
# student={
#     "name":"shubham",
#     "id" :20,
#     "address" :"umaria"
# }
# student["dob"]="21/12/1989"
# print(student)
# print(student.get("id"))
digit_to_word = {
    "1": "one",
    "2": "two",
    "3": "three",
    "4": "four",
    "5": "five",
    "6": "six",
    "7": "seven",
    "8": "eight",
    "9": "nine",
    "0": "zero"
}

# phone_number=input("enter the phone # ")
#
# phone_in_word=''
#
# for digit in phone_number:
#     phone_in_word+=digit_to_word.get(digit)+' '
#
# print(phone_in_word)


# emoji
#
# emoji={
#     ":)":"☺",
#     ":(": "😥"
#
# }
# words=input("Greeeting!!>")
# words=words.split(' ')
# return_val=''
# for word in words:
#     return_val+=  emoji.get(word,word)+' '
# print(return_val)
#
# def greeting(first_name,last_name):
#     print(f'Hi {first_name} {last_name}')
#
# print("start")
# greeting("shubham",last_name="Tripathi")


# exception handling
#
# try:
#     age = int(input("Age : "))
#     salary=20000
#     op=salary/age
#     print(age)
# except ZeroDivisionError:
#     print('age not be 0')
# except ValueError:
#     print("invalid number")

# classes in python
#
# class Point:
#     def __init__(self, x, y):
#         self.x = x
#         self.y = y
#
#     def draw(self):
#         print("draw",self.x)
#
#
# point = Point(10,20)
# point.draw()
# print(point.x)


# inheritance in python
#
# class Mammal:
#     def walk(self):
#         print("walk")
#
#
# class Dog(Mammal):
#     def bark(self):
#         print("bark")
#
#
# class Cat(Mammal):
#     pass
#
#
# dog = Dog()
# dog.walk()
#
# cat = Cat()
# cat.walk()


# modules
# import converters
#
# print(converters.find_max(numbers=[10,23,54,99]))


# package
#
# import ecommerce.shipping
# ecommerce.shipping.calculate_cost()
#
# from ecommerce.shipping import calculate_cost
# calculate_cost()


# random module

# import random
# for item in range(3):
#     print(random.random())
#     print(random.randint(10,30))
#     leaders=["MASH", "SHUBHAM", "BOB"]
#     print(random.choice(leaders))
#
#
# import random
#
#
# class Dice:
#     def roll(self):
#         dice1 = random.randint(1, 6)
#         dice2 = random.randint(1, 6)
#         tuple_arr = (dice1, dice2)
#         return tuple_arr
#
#
# dice = Dice()
# print(dice.roll())


# WORK WITH FILE AND PATH

#
# from pathlib import Path
#
# path = Path('ecommerce')
# print(path.exists())
# path = Path('emails')
# if not path.exists():
#     path.mkdir()
# print(path.exists())
# path.rmdir()
# for file in path.glob("*.*"):
#     print(file.name)

import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
    workbook = xl.load_workbook(filename)
    sheet1 = workbook['Sheet1']
    for row in range(2, sheet1.max_row + 1):
        actual_price = sheet1.cell(row, 3)
        discounted_price = actual_price.value * .9
        discounted_cell = sheet1.cell(row, 4)
        discounted_cell.value = discounted_price

    ref_discounted_values = Reference(sheet1, min_row=2, max_row=sheet1.max_row, min_col=4, max_col=4)
    chart = BarChart()
    chart.add_data(ref_discounted_values)
    sheet1.add_chart(chart, 'e2')
    workbook.save(filename)


process_workbook("transactions.xlsx")